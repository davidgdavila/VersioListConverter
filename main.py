import threading
import os
import time
import tkinter as tk
from tkinter import filedialog as fd
from kivy import Config
from kivy.core.window import Window
from kivy.lang import Builder
from kivy.properties import StringProperty, AliasProperty, NumericProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from openpyxl import load_workbook


def seleccionar_salvar_archivo(ruta_abrir):
    root = tk.Tk()
    root.withdraw()
    d = os.path.dirname(ruta_abrir)
    s = os.path.basename(ruta_abrir).replace(".xlsx", "")
    filename = fd.asksaveasfilename(
        initialfile="versio_" + s + ".txt",
        title='Salvar Archivo',
        initialdir=d,
        filetypes=(('text files', '*.txt'), ('All files', '*.*')))
    ruta_guardar = filename
    if filename:
        hilo1 = threading.Thread(target=procesar_archivo, args=(ruta_abrir, ruta_guardar))
        hilo1.start()
    else:
        MDApp.get_running_app().root.eti1.text = "Arrastra la pauta aqui."

def cargar_barra():
    for i in range(0,100):
        valor =+ i
        time.sleep(0.03)
        MDApp.get_running_app().root.eti1.text = "Cargando archivo..." + str(round(valor)) + "%"
        MDApp.get_running_app().root.barra.value = valor

def procesar_archivo(ruta_abrir,ruta_guardar):
    texto = ""
    try:
        hilo2 = threading.Thread(target=cargar_barra)
        hilo2.start()
        s = os.path.basename(ruta_abrir)
        #MDApp.get_running_app().root.eti1.text = "Cargando " + s + "..."
        doc = load_workbook(filename=ruta_abrir, data_only=True)
        #sheet1 = doc.sheetnames[0]
        #hoja1 = doc[sheet1]
        hoja1=doc.active
        filas = tuple(hoja1.rows)
        longitudfilas = len(filas)
        MDApp.get_running_app().root.eti1.text = "Convirtiendo archivo..."
        for i in range(5, longitudfilas):
            E = hoja1["E"+ str(i)].value
            F = hoja1["F" + str(i)].value
            G = hoja1["G" + str(i)].value
            H = hoja1["H" + str(i)].value
            J = hoja1["J" + str(i)].value
            L = hoja1["L" + str(i)].value
            texto += "\t\t\t\t"+str(E) + "\t" + str(F) + "\t" + str(G) + "\t"\
                     + str(H) + "\t\t" + str(J) + "\t\t" + str(L) + "\n"
            texto = texto.replace(str(None), "")

            #valor = i*100/longitudfilas
            #MDApp.get_running_app().root.eti1.text = "Convirtiendo archivo..." + str(round(valor))+"%"
            #MDApp.get_running_app().root.barra.value = valor

        d =ruta_guardar.replace(".txt","")
        with open(d +".txt", 'w') as stream:
            stream.write(texto)
        MDApp.get_running_app().root.eti1.text = "¡Se ha convertido el archivo!\n" + ruta_guardar
    except IndexError:
        MDApp.get_running_app().root.eti1.text = "El formato no es correcto"
    except UnicodeDecodeError:
        MDApp.get_running_app().root.eti1.text = "Error:¡Checa que sea un txt wey!"
    except:
        MDApp.get_running_app().root.eti1.text = "No se realizó la conversion.Error desconocido"

class VentanaPrincipal(BoxLayout):
    texto = ""
    ruta_abrir = ""
    ruta_guardar=""
    nombre_archivo = ""
    etiqueta = StringProperty("Arrastra la pauta aqui.")
    directorio_inicial = os.path.expanduser("~") + "\\documents\\"
    longitudfilas= NumericProperty(0)

    def __init__(self,**kwargs):
        super().__init__(**kwargs)
        Window.bind(on_drop_file=self._on_file_drop)

    def _on_file_drop(self, window, file_path, *args):
        MDApp.get_running_app().root.eti1.text = "Elija la pauta que desea convertir..."
        self.ruta_abrir = file_path.decode(encoding="utf-8")
        hilo3 = threading.Thread(target=seleccionar_salvar_archivo, args=(self.ruta_abrir,))
        hilo3.start()

    def seleccionar_archivo(self):
        root = tk.Tk()
        root.withdraw()
        filename = fd.askopenfilename(
            title='Seleccionar Pauta...',
            initialdir=self.directorio_inicial,
            filetypes=(('excel files', '*.xlsx'), ('All files', '*.*')))
        self.ruta_abrir = filename
        if filename:
            self.seleccionar_salvar_archivo()
        else:
            pass

    def seleccionar_salvar_archivo(self):
        root = tk.Tk()
        root.withdraw()
        d=os.path.dirname(self.ruta_abrir)
        s=os.path.basename(self.ruta_abrir).replace(".xlsx","")
        filename = fd.asksaveasfilename(
            initialfile= "versio_"+s+".txt",
            title='Salvar Archivo',
            initialdir=d,
            filetypes=(('text files', '*.txt'), ('All files', '*.*')))
        self.ruta_guardar=filename
        if filename:
            hilo1 = threading.Thread(target=procesar_archivo, args=(self.ruta_abrir,self.ruta_guardar))
            hilo1.start()
        else:
            pass


    """ def guardar_archivo(self):
        if ".txt" in self.ruta_guardar:
            self.ruta_guardar.replace(".txt","")
        with open(self.ruta_guardar, 'w') as stream:
            stream.write(self.texto)
        self.etiqueta = "¡Se ha convertido el archivo!\n" + self.ruta_guardar"""

    def show_acerca_de(self):
        contenido = Acerca_de()
        self._popup = Popup(
            title='Acerca de Versio List Converter',
            content=contenido,
            size_hint=(None, None), size=(400, 400))
        self._popup.open()

    def limpiar(self):
        self.ids.eti1.text="Arrastra el archivo aqui."
        self.ids.barra.value=0

class Acerca_de(MDBoxLayout):
    texto_etiqueta = StringProperty("Autor: David Israel González Dávila.\n\n"
                                "Versión: 1.1\n\n"
                                "Fecha: 12/abril/2022\n\n"
                                "Descripción: Generar el formato correcto para\n"
                                "la elaboración de la playlist.\n"
                                "Objetivo: Elaborar la playlist de una forma\n"
                                "más eficiente\n")

class VersioListConverterApp(MDApp):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        Config.set("kivy", "window_icon", "versiolist.ico")
        self.title = "VersioListConverter"
        self.icon = "versiolist.ico"

    def build(self):
        self.root = Builder.load_file("Versio_List_Converter.kv")
VersioListConverterApp().run()
